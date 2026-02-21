# frontend/streamlit_app.py
import streamlit as st
import requests
from datetime import date
import logging
import sys
import re
from streamlit_pdf_viewer import pdf_viewer
from pathlib import Path
import io
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
import fitz
import json
import os
import comtypes.client


BACKEND_URL = "http://localhost:8001"
json_file = Path(__file__).parent / "result_temp.json"

if "preview_id" not in st.session_state:
    st.session_state["preview_id"] = None

if "sem_preview_id" not in st.session_state:
    st.session_state["sem_preview_id"] = None

if "hybrid_preview_id" not in st.session_state:
    st.session_state["hybrid_preview_id"] = None
                    
if "data" not in st.session_state:
    st.session_state["data"] = {}

if 'initialized' not in st.session_state:
    if os.path.exists(json_file):
        os.remove(json_file)
    st.session_state['initialized'] = True

if "search_type" not in st.session_state:
    st.session_state["search_type"] = None  # default to keyword search


# Configure basic server-side logging: writes to stdout (visible in the terminal
# running Streamlit) and to a file `streamlit_server.log` in the workspace.
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("streamlit_server.log", encoding="utf-8"),
    ],
)

# Utility function to save search results to a JSON file
def save_results_to_json(results):
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=4)

st.set_page_config(page_title="Board Smart Search", layout="wide")
st.title("Board Smart Search – Prototype")

st.markdown(
    """
    <style>
    div[data-testid="stButton"] > button[kind="tertiary"],
    div[data-testid="stButton"] > button[data-testid="stBaseButton-tertiary"] {
        justify-content: flex-start !important;
        padding-left: 0.1rem !important;
    }

    div[data-testid="stButton"] > button[kind="tertiary"] p,
    div[data-testid="stButton"] > button[kind="tertiary"] span,
    div[data-testid="stButton"] > button[kind="tertiary"] div,
    div[data-testid="stButton"] > button[data-testid="stBaseButton-tertiary"] p,
    div[data-testid="stButton"] > button[data-testid="stBaseButton-tertiary"] span,
    div[data-testid="stButton"] > button[data-testid="stBaseButton-tertiary"] div {
        color: #1f77b4 !important;
        text-decoration: underline !important;
        font-weight: 700 !important;
        font-size: 20px !important;
        line-height: 1.35 !important;
        margin: 0 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# Function to highlight exact keyword matches in the text
def highlight_exact_query(text, query):
    
    # Use word boundaries (\b) in regex to match the exact word
    # The pattern is a raw string r'\bquery\b' to ensure correct interpretation of \b
    pattern = r'\b' + re.escape(query) + r'(?:\'s|es|s)?\b'
    # Replace the exact matches with the colored version
    highlighted_text = re.sub(pattern, r":blue-background[\g<0>]", text, flags=re.IGNORECASE)
    
    return highlighted_text

# Function to handle PDF highlighting based on search results
def highlight_pdf(highlight_payload: dict, target_path):
    doc = fitz.open(target_path)
    if not isinstance(highlight_payload, dict):
        raise TypeError("highlight_payload must be a dict with keys: mode, targets")

    mode = highlight_payload.get("mode", "line")
    targets = highlight_payload.get("targets", [])
    first_match_page = None
            
    # Iterate through each page to find and highlight string
    for page_num in range(len(doc)):
        page = doc[page_num]
        if mode == "keyword":
            matched_targets = targets
        else:
            page_lines = _get_page_text_lines(page)
            matched_targets = _match_exact_lines(targets, page_lines)

        for target in matched_targets:
            rect_list = page.search_for(
                target,
                flags=fitz.TEXT_DEHYPHENATE | fitz.TEXT_PRESERVE_WHITESPACE | fitz.TEXT_MEDIABOX_CLIP,
            )
            if rect_list and first_match_page is None:
                first_match_page = page_num + 1
            for inst in rect_list:
                page.add_highlight_annot(inst).update()

              
    # Save the modified PDF to a memory buffer (raw bytes)
    pdf_bytes = doc.tobytes() 
    doc.close()
    return pdf_bytes, first_match_page

# Convert snippets/query text into lines for better matching in previews
def _build_search_lines(keywords) -> list[str]:
    if keywords is None:
        return []

    if isinstance(keywords, (list, tuple, set)):
        chunks = [str(k) for k in keywords if str(k).strip()]
    else:
        chunks = [str(keywords)]

    lines = []
    for chunk in chunks:
        cleaned_chunk = str(chunk).replace("**", "")
        normalized_chunk = re.sub(r"\s+", " ", cleaned_chunk).strip()
        if not normalized_chunk:
            continue

        for line in normalized_chunk.splitlines():
            cleaned_line = re.sub(r"\s+", " ", line).strip()
            if cleaned_line:
                lines.append(cleaned_line)

        sentence_parts = re.split(r"(?<=[.!?])\s+", normalized_chunk)
        for sentence in sentence_parts:
            cleaned_sentence = re.sub(r"\s+", " ", sentence).strip()
            if len(cleaned_sentence) >= 20:
                lines.append(cleaned_sentence)

    return list(dict.fromkeys(lines))

# Extract keyword term from query for keyword-based highlighting in previews
def _build_keyword_terms(query_text) -> list[str]:
    if query_text is None:
        return []

    if isinstance(query_text, (list, tuple, set)):
        text = " ".join(str(item) for item in query_text)
    else:
        text = str(query_text)

    terms = [term for term in re.findall(r"[A-Za-z0-9']+", text.lower()) if len(term) >= 2]
    return list(dict.fromkeys(terms))

# Function to normalize lines for PDF,XLSX and PPT in semantic/hybrid search
def _normalize_line_for_match(line: str) -> str:
    line = re.sub(r"\s+", " ", line).strip().lower()
    line = re.sub(r"^[^\w]+|[^\w]+$", "", line)
    return line

# Extract text lines from a PDF page
def _get_page_text_lines(page) -> list[str]:
    text = page.get_text("text") or ""
    lines = []
    for line in text.splitlines():
        cleaned = re.sub(r"\s+", " ", line).strip()
        if cleaned:
            lines.append(cleaned)
    return lines

# Match lines from the PDF page to the target lines for highlighting in semantic search previews
def _match_exact_lines(search_lines: list[str], page_lines: list[str]) -> list[str]:
    wanted = [_normalize_line_for_match(line) for line in search_lines if line]
    matched = []
    seen = set()
    for page_line in page_lines:
        normalized = _normalize_line_for_match(page_line)
        if not normalized:
            continue

        line_match = False
        for wanted_line in wanted:
            if not wanted_line:
                continue
            if normalized == wanted_line:
                line_match = True
                break
            if len(normalized) >= 20 and (normalized in wanted_line or wanted_line in normalized):
                line_match = True
                break

        if line_match and page_line not in seen:
            matched.append(page_line)
            seen.add(page_line)
    return matched

# For excel, semantic search preview to match snippet lines to cell values
def _text_matches_line_targets(text: str, targets: list[str]) -> bool:
    normalized_text = _normalize_line_for_match(text)
    if not normalized_text:
        return False

    for target in targets:
        normalized_target = _normalize_line_for_match(target)
        if not normalized_target:
            continue
        if normalized_text == normalized_target:
            return True
        if len(normalized_text) >= 20 and (normalized_text in normalized_target or normalized_target in normalized_text):
            return True

    return False


# Highlighting for PPTX files: converts to PDF, 
# applies similar highlighting as for PDFs, 
# then returns PDF bytes for preview rendering
def highlight_pptx(highlight_payload: dict, target_path):  
    # 1. Convert PPT to PDF using comtypes
    if not isinstance(highlight_payload, dict):
        raise TypeError("highlight_payload must be a dict with keys: mode, targets")

    mode = highlight_payload.get("mode", "line")
    targets = highlight_payload.get("targets", [])
    first_match_page = None
        
    ppt_path = os.path.abspath(target_path)
    pdf_path = ppt_path.rsplit(".", 1)[0] + ".pdf"
    pdf_path = os.path.abspath(pdf_path)
    try:
        comtypes.CoInitialize() 
        powerpoint = comtypes.client.CreateObject("PowerPoint.Application")
        powerpoint.Visible = 1

        # Open and Export
        deck = powerpoint.Presentations.Open(ppt_path)
        # Format 32 is PDF
        deck.ExportAsFixedFormat(pdf_path, 32)
        deck.Close()
        powerpoint.Quit()
    except Exception as e:
        logging.error(f"Failed to convert PPT to PDF: {e}")
        return None
    finally:
        comtypes.CoUninitialize()

    # 2. Highlight and Fetch Bytes
    doc = fitz.open(pdf_path)
    for page_num, page in enumerate(doc):
        total_matches = 0
        if mode == "keyword":
            matched_targets = targets
        else:
            page_lines = _get_page_text_lines(page)
            matched_targets = _match_exact_lines(targets, page_lines)

        for target in matched_targets:
            rect_list = page.search_for(
                target,
                flags=fitz.TEXT_DEHYPHENATE | fitz.TEXT_PRESERVE_WHITESPACE | fitz.TEXT_MEDIABOX_CLIP,
            )
            if rect_list and first_match_page is None:
                first_match_page = page_num + 1
            for inst in rect_list:
                page.add_highlight_annot(inst).update()
            total_matches += len(rect_list)
        logging.info(f"Found {total_matches} occurrences on page {page.number + 1}")
    
    # Get PDF data as bytes directly from memory
    pdf_bytes = doc.tobytes() 
    doc.close()

    # 3. Cleanup: Remove the temporary PDF file
    if os.path.exists(pdf_path):
        os.remove(pdf_path)
    
    return pdf_bytes, first_match_page

# Highlighting for DOCX files: converts to PDF,
# applies similar highlighting as for PDFs,
# then returns PDF bytes for preview rendering
def highlight_docx(highlight_payload: dict, target_path):
    # 1. Convert DOCX to PDF using comtypes (Word)
    if not isinstance(highlight_payload, dict):
        raise TypeError("highlight_payload must be a dict with keys: mode, targets")

    mode = highlight_payload.get("mode", "line")
    targets = highlight_payload.get("targets", [])
    first_match_page = None
        
    docx_path = os.path.abspath(str(target_path))
    pdf_path = docx_path.rsplit(".", 1)[0] + ".pdf"
    pdf_path = os.path.abspath(pdf_path)
    try:
        comtypes.CoInitialize()
        word = comtypes.client.CreateObject("Word.Application")
        word.Visible = 0  # Run Word in background

        # Open the DOCX document
        doc = word.Documents.Open(docx_path)
        # ExportAsFixedFormat: 17 = wdExportFormatPDF
        doc.ExportAsFixedFormat(pdf_path, 17)
        doc.Close()
        word.Quit()
    except Exception as e:
        logging.error(f"Failed to convert DOCX to PDF: {e}")
        return None, None
    finally:
        comtypes.CoUninitialize()

    # 2. Highlight and Fetch Bytes
    doc = fitz.open(pdf_path)
    for page_num, page in enumerate(doc):
        total_matches = 0
        if mode == "keyword":
            matched_targets = targets
        else:
            page_lines = _get_page_text_lines(page)
            matched_targets = _match_exact_lines(targets, page_lines)

        for target in matched_targets:
            rect_list = page.search_for(
                target,
                flags=fitz.TEXT_DEHYPHENATE | fitz.TEXT_PRESERVE_WHITESPACE | fitz.TEXT_MEDIABOX_CLIP,
            )
            if rect_list and first_match_page is None:
                first_match_page = page_num + 1
            for inst in rect_list:
                page.add_highlight_annot(inst).update()
            total_matches += len(rect_list)
        logging.info(f"Found {total_matches} occurrences on page {page.number + 1}")
    
    # Get PDF data as bytes directly from memory
    pdf_bytes = doc.tobytes()
    doc.close()

    # 3. Cleanup: Remove the temporary PDF file
    if os.path.exists(pdf_path):
        os.remove(pdf_path)
    
    return pdf_bytes, first_match_page

# Handler for preview button clicks, 
# sets the appropriate preview ID in session state 
# and prepares highlighted data for rendering
def set_preview_id(document_id, file_path, idx, keywords, page_start=None):
    if st.session_state["search_type"] == "keyword":
        st.session_state["preview_id"] = idx  # Store the result index, not document_id
    elif st.session_state["search_type"] == "semantic":
        st.session_state["sem_preview_id"] = idx
    elif st.session_state["search_type"] == "hybrid":
        st.session_state["hybrid_preview_id"] = idx
    
    logging.info(f"file_path={file_path}, idx={idx}, page_start={page_start}")
        
    target_path = Path(file_path)

    if st.session_state.get("search_type") == "keyword":
        highlight_mode = "keyword"
        highlight_targets = _build_keyword_terms(keywords)
    else:
        highlight_mode = "line"
        highlight_targets = _build_search_lines(keywords)

    highlight_payload = {"mode": highlight_mode, "targets": highlight_targets}

    if target_path.exists():
        if file_path.lower().endswith(".pdf"):
            logging.info(f"Preparing PDF preview for {file_path}")
            # Open the PDF using PyMuPDF (fitz)
            pdf_bytes, first_match_page = highlight_pdf(highlight_payload, target_path)
            st.session_state[f"data_{idx}"] = pdf_bytes
            # Use page_start from search result if available, otherwise use first_match_page from highlighting
            st.session_state[f"first_match_page_{idx}"] = page_start or first_match_page or 1
        elif file_path.lower().endswith((".doc", ".docx")):
            logging.info(f"Preparing Word document preview for {file_path}")
            # Convert DOCX to PDF and apply highlighting
            pdf_bytes, first_match_page = highlight_docx(highlight_payload, target_path)
            st.session_state[f"data_{idx}"] = pdf_bytes
            # Use page_start from search result if available, otherwise use first_match_page from highlighting
            st.session_state[f"first_match_page_{idx}"] = page_start or first_match_page or 1
        elif file_path.lower().endswith((".xls", ".xlsx")):
            # for excel, highlight excel with keywords by reading the file and converting to html with highlights
            logging.info(f"Preparing Excel preview for {file_path}")
            with open(target_path, "rb") as f:
                st.session_state[f"data_{idx}"] = f.read()
                wb = None
                try:
                    wb = load_workbook(filename=io.BytesIO(st.session_state[f"data_{idx}"]), read_only=True)
                    frames = []
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        data = list(sheet.values)
                        if not data:
                            continue
                        sheet_df = pd.DataFrame(data)
                        if sheet_df.empty:
                            continue
                        sheet_df.insert(0, "Sheet", sheet_name)
                        frames.append(sheet_df)

                    if not frames:
                        st.session_state[f"data_{idx}"] = pd.DataFrame({"Sheet": [], "Value": []})
                    else:
                        df = pd.concat(frames, ignore_index=True)

                        keyword_terms = highlight_targets

                        if keyword_terms:
                            keyword_pattern = None
                            if highlight_mode == "keyword":
                                keyword_pattern = r"\b(" + "|".join(map(re.escape, keyword_terms)) + r")\b"

                            def apply_highlight(val):
                                if val is None:
                                    return ''

                                text_val = val if isinstance(val, str) else str(val)
                                if highlight_mode == "keyword":
                                    if keyword_pattern and re.search(keyword_pattern, text_val, flags=re.IGNORECASE):
                                        return 'background-color: yellow; color: black;'
                                else:
                                    if _text_matches_line_targets(text_val, keyword_terms):
                                        return 'background-color: yellow; color: black;'
                                return ''

                            # Create the styled object
                            styled_df = df.style.map(apply_highlight)
                            st.session_state[f"data_{idx}"] = styled_df
                        else:
                            # No keywords, store plain dataframe
                            st.session_state[f"data_{idx}"] = df

                except Exception as e:
                    logging.error("Failed to process Excel file for preview: %s", e)
                    st.session_state[f"data_{idx}"] = None
                finally:
                    if wb is not None:
                        wb.close()
        elif file_path.lower().endswith((".pptx",".ppt")):
            pdf_bytes, first_match_page = highlight_pptx(highlight_payload, target_path)
            st.session_state[f"data_{idx}"] = pdf_bytes
            # Use page_start from search result if available, otherwise use first_match_page from highlighting
            st.session_state[f"first_match_page_{idx}"] = page_start or first_match_page or 1
        else:
            st.warning("Unsupported file type for preview.")

tab_search, tab_admin, tab_upload = st.tabs(
    ["🔍 Search (stub)", "🏗️ Committees & Meetings", "📄 Upload Documents"]
)

with tab_search:
    col1, col2 = st.columns([0.6, 0.4])
    with col1:
        st.subheader("Search")

        committee_external_id = st.text_input("Committee External ID", "AC001")

        meetings_options = ["All meetings"]
        meeting_map = {}

        if committee_external_id:
            try:
                resp = requests.get(
                    f"{BACKEND_URL}/meetings/",
                    params={"committee_external_id": committee_external_id},
                    timeout=5,
                )
                if resp.status_code == 200:
                    meetings = resp.json()
                    for m in meetings:
                        label = f"{m['external_meeting_id']} – {m['name']}"
                        meetings_options.append(label)
                        meeting_map[label] = m["external_meeting_id"]
                else:
                    st.warning(
                        f"Could not load meetings for committee {committee_external_id} "
                        f"(status {resp.status_code})."
                    )
            except Exception as e:
                st.warning(f"Error loading meetings: {e}")

        selected_meeting_label = st.selectbox("Meeting (optional filter)", meetings_options)
        meeting_external_id = (
            None if selected_meeting_label == "All meetings" else meeting_map[selected_meeting_label]
        )

        doc_type_filter = st.selectbox(
            "Document Type (optional filter)",
            ["Any", "Agenda", "DraftMinutes", "FinalMinutes", "CircularResolution", "Extra1", "Extra2"],
        )
        doc_type_value = None if doc_type_filter == "Any" else doc_type_filter

        
        st.subheader("Keyword Search")

        query = st.text_input("Search query", key="keyword_query")

        if st.button("Search", type="primary"):
            if json_file.exists():
                json_file.unlink()  # Clear previous results

            st.session_state["preview_id"] = None
                
            payload = {
                "committee_external_id": committee_external_id,
                "query": query,
                "meeting_external_id": meeting_external_id,
                "doc_type": doc_type_value,
            }
            try:
                resp = requests.post(f"{BACKEND_URL}/search/", json=payload)
                st.session_state["search_type"] = "keyword"
                if resp.status_code != 200:
                    st.error(f"Error {resp.status_code}: {resp.text}")
                    st.session_state["last_results"] = []
                    st.session_state["last_query"] = query
                else:
                    data = resp.json()
                    results = data.get("results", [])
                    st.session_state["last_results"] = results
                    save_results_to_json(results)
                    st.session_state["last_query"] = query
                    if not results:
                        st.info("No results found.")
                        st.session_state["preview_id"] = None
                    else:
                        st.success(f"Found {len(results)} result(s).")
                        first_hit = results[0]
                        first_file_path = first_hit.get("file_path")
                        first_keywords = query if query else None
                        first_page_start = first_hit.get("page_start")
                        set_preview_id(first_hit["document_id"], first_file_path, 0, first_keywords, first_page_start)
            except Exception as e:
                st.error(f"Search failed: {e}")

        if st.session_state.get("search_type") == "keyword":
            # Render last search results from session state so previews work on rerun
            results = json_file.read_text(encoding="utf-8") if json_file.exists() else []
            results = json.loads(results) if results else []

            last_query = str(st.session_state.get("last_query", ""))

            if results:
                for idx, hit in enumerate(results):
                    file_path = hit.get("file_path")
                    doc_id = hit["document_id"]
                    keywords = last_query if last_query else None
                    page_start = hit.get("page_start")

                    #st.subheader(hit['document_title'])
                    
                    st.button(
                        f"**{hit['document_title']}**",
                        key=f"doc_title_{idx}",
                        type="tertiary",
                        on_click=set_preview_id,
                        args=(doc_id, file_path, idx, keywords, page_start),
                    )

                    meta = f"Doc type: {hit['doc_type']}"
                    if hit.get("meeting_name"):
                        meta += f" | Meeting: {hit['meeting_name']}"
                        meta += f" | Page: {page_start}"
                    st.caption(meta)

                    highlighted_snippet = highlight_exact_query(hit["snippet"], keywords)
                    st.markdown(highlighted_snippet)

                    st.markdown(
                        f"*Matches in this section:* `{hit['occurrence_count']}`  "
                        f"| *Score:* `{hit['score']:.3f}`"
                    )
                    if hit.get("ddm_url"):
                        st.markdown(f"[Open in DDM]({hit['ddm_url']})")
                    
        
        st.markdown("---")
        st.subheader("Semantic Search (beta)")

        sem_query = st.text_input(
            "Semantic query (concepts, synonyms, etc.)",
            key="semantic_query",
            placeholder="e.g. related party transactions, board evaluation, ESOP approvals",
        )

        if st.button("Semantic Search", type="secondary"):
            if json_file.exists():
                json_file.unlink()  # Clear previous results
            st.session_state["sem_preview_id"] = None
            payload = {
                "committee_external_id": committee_external_id,
                "query": sem_query,
                "meeting_external_id": meeting_external_id,
                "doc_type": doc_type_value,
            }
            try:
                resp = requests.post(f"{BACKEND_URL}/semantic-search/", json=payload)
                st.session_state["search_type"] = "semantic"
                if resp.status_code != 200:
                    st.error(f"Error {resp.status_code}: {resp.text}")
                    st.session_state["last_semantic_results"] = []
                    st.session_state["last_semantic_query"] = sem_query
                else:
                    data = resp.json()
                    results = data.get("results", [])
                    st.session_state["last_semantic_results"] = results
                    save_results_to_json(results)
                    st.session_state["last_semantic_query"] = sem_query 
                    if not results:
                        st.info("No semantic results found.")
                        st.session_state["sem_preview_id"] = None
                    else:
                        st.success(f"Found {len(results)} semantic result(s).")
                        first_hit = results[0]
                        first_file_path = first_hit.get("file_path")
                        first_keywords = str(first_hit.get("snippet")) if first_hit.get("snippet") else (sem_query if sem_query else None)
                        first_page_start = first_hit.get("page_start")
                        set_preview_id(first_hit["document_id"], first_file_path, 0, first_keywords, first_page_start)
            except Exception as e:
                st.error(f"Semantic search failed: {e}")

        if st.session_state.get("search_type") == "semantic":
            sem_results = json_file.read_text(encoding="utf-8") if json_file.exists() else []
            sem_results = json.loads(sem_results) if sem_results else []

            if sem_results:
                st.subheader("Semantic Search Results")
                for idx, hit in enumerate(sem_results):
                    file_path = hit.get("file_path")
                    doc_id = hit["document_id"]
                    keywords = str(hit["snippet"]) if hit.get("snippet") else None
                    page_start = hit.get("page_start")

                    st.button(
                        f"**{hit['document_title']}**",
                        key=f"doc_title_sem_{idx}",
                        type="tertiary",
                        on_click=set_preview_id,
                        args=(doc_id, file_path, idx, keywords, page_start),
                    )

                    meta_parts = []
                    if hit.get("meeting_name"):
                        meta_parts.append(f"Meeting: {hit['meeting_name']}")
                    meta_parts.append(f"Doc type: {hit['doc_type']}")
                    st.caption(" | ".join(meta_parts))

                    snippet = hit["snippet"]

                    st.write(snippet)

                    st.markdown(f"*Semantic score:* `{hit['score']:.3f}`")
                    if hit.get("ddm_url"):
                        st.markdown(f"[Open in DDM]({hit['ddm_url']})")
                    st.markdown("---")

    
        st.markdown("---")
        st.subheader("Hybrid Search (recommended)")

        hybrid_query = st.text_input(
            "Hybrid query (uses both keyword + semantic signals)",
            key="hybrid_query",
            placeholder="e.g. appropriation of profit as per statutory requirement",
        )

        if st.button("Hybrid Search"):
            if json_file.exists():
                json_file.unlink()  # Clear previous results
            st.session_state["hybrid_preview_id"] = None
            payload = {
                "committee_external_id": committee_external_id,
                "query": hybrid_query,
                "meeting_external_id": meeting_external_id,
                "doc_type": doc_type_value,
            }
            try:
                resp = requests.post(f"{BACKEND_URL}/hybrid-search/", json=payload)
                st.session_state["search_type"] = "hybrid"
                if resp.status_code != 200:
                    st.error(f"Error {resp.status_code}: {resp.text}")
                    st.session_state["last_hybrid_results"] = []
                    st.session_state["last_hybrid_query"] = hybrid_query
                else:
                    data = resp.json()
                    results = data.get("results", [])
                    st.session_state["last_hybrid_results"] = results
                    save_results_to_json(results)
                    st.session_state["last_hybrid_query"] = hybrid_query
                    if not results:
                        st.info("No hybrid results found.")
                        st.session_state["hybrid_preview_id"] = None
                    else:
                        st.success(f"Found {len(results)} hybrid result(s). (sorted by score)")
                        sorted_results = sorted(results, key=lambda h: h["score"], reverse=True)
                        save_results_to_json(sorted_results)
                        st.session_state["last_hybrid_results"] = sorted_results
                        first_hit = sorted_results[0]
                        first_file_path = first_hit.get("file_path")
                        first_keywords = str(first_hit.get("snippet")) if first_hit.get("snippet") else (hybrid_query if hybrid_query else None)
                        first_page_start = first_hit.get("page_start")
                        set_preview_id(first_hit["document_id"], first_file_path, 0, first_keywords, first_page_start)
            except Exception as e:
                st.error(f"Hybrid search failed: {e}")
        
        if st.session_state.get("search_type") == "hybrid":
            hybrid_results = json_file.read_text(encoding="utf-8") if json_file.exists() else []
            hybrid_results = json.loads(hybrid_results) if hybrid_results else []

            if hybrid_results:
                st.subheader("Hybrid Search Results")
                results = sorted(hybrid_results, key=lambda h: h["score"], reverse=True)
                st.success(f"Found {len(results)} hybrid result(s). (sorted by score)")

                for idx, hit in enumerate(results):
                    file_path = hit.get("file_path")
                    doc_id = hit["document_id"]
                    keywords = str(hit["snippet"]) if hit.get("snippet") else None
                    page_start = hit.get("page_start")

                    st.button(
                        f"**{hit['document_title']}**",
                        key=f"doc_title_hybrid_{idx}",
                        type="tertiary",
                        on_click=set_preview_id,
                        args=(doc_id, file_path, idx, keywords, page_start),
                    )

                    meta_parts = []
                    if hit.get("meeting_name"):
                        meta_parts.append(f"Meeting: {hit['meeting_name']}")
                    meta_parts.append(f"Doc type: {hit['doc_type']}")
                    st.caption(" | ".join(meta_parts))

                    snippet = hit["snippet"]

                    st.write(snippet)

                    st.markdown(f"*Hybrid score:* `{hit['score']:.3f}`")
                    if hit.get("ddm_url"):
                        st.markdown(f"[Open in DDM]({hit['ddm_url']})")
                    st.markdown("---")
    with col2:
        st.subheader("Preview Document")

        current_results = json_file.read_text(encoding="utf-8") if json_file.exists() else []
        current_results = json.loads(current_results) if current_results else []

        current_search_type = st.session_state.get("search_type")
        preview_key_map = {
            "keyword": "preview_id",
            "semantic": "sem_preview_id",
            "hybrid": "hybrid_preview_id",
        }
        preview_state_key = preview_key_map.get(current_search_type)
        preview_idx = st.session_state.get(preview_state_key) if preview_state_key else None

        can_preview = (
            preview_state_key is not None
            and isinstance(preview_idx, int)
            and 0 <= preview_idx < len(current_results)
        )

        if can_preview:
            hit = current_results[preview_idx]
            idx = preview_idx
            file_path = hit.get("file_path")

            if not file_path:
                st.info("No file is available for preview on the selected result.")
            else:
                with st.container(border=True):
                    st.write(f"🔍 **Previewing:** {hit['document_title']}")

                    binary_data = st.session_state.get(f"data_{idx}")
                    if st.button("Close Preview ✖️", key=f"close_preview_{current_search_type}_{idx}"):
                        st.session_state[preview_state_key] = None
                        st.rerun()

                    try:
                        if file_path.lower().endswith('.pdf'):
                            st.write("Rendering PDF preview...")
                            first_page = st.session_state.get(f"first_match_page_{idx}", 1)
                            target_page = first_page
                            pdf_viewer(input=binary_data, height=700, width="100%", key=f"pdf_preview_{current_search_type}_{idx}_{target_page}", scroll_to_page=target_page)
                        elif file_path.lower().endswith((".doc", ".docx")):
                            st.write("Rendering Word preview...")
                            first_page = st.session_state.get(f"first_match_page_{idx}", 1)
                            target_page = first_page 
                            pdf_viewer(input=binary_data, height=700, width="100%", key=f"docx_preview_{current_search_type}_{idx}_{target_page}", scroll_to_page=target_page)
                        elif file_path.lower().endswith((".xls", ".xlsx")):
                            st.write("Rendering Excel preview...")
                            st.dataframe(binary_data, key=f"excel_preview_{current_search_type}_{idx}", use_container_width=True)
                        elif file_path.lower().endswith((".pptx", ".ppt")):
                            st.write("Rendering ppt Preview...")
                            first_page = st.session_state.get(f"first_match_page_{idx}", 1)
                            target_page = first_page
                            pdf_viewer(input=binary_data, height=700, width="100%", key=f"ppt_preview_{current_search_type}_{idx}_{target_page}", scroll_to_page=target_page)
                        else:
                            st.warning("Unsupported file type for preview.")
                    except Exception as e:
                        st.error(f"Failed to render preview: {e}")
        elif current_search_type in {"keyword", "semantic", "hybrid"}:
            st.info("Run a search and select a result to preview.")




with tab_admin:
    st.subheader("Create Committee")

    col1, col2 = st.columns(2)

    with col1:
        company_external_id = st.text_input(
            "Company External ID",
            value="Titan",
            help="E.g. Titan; used as tenant/company key",
        )
        committee_name = st.text_input(
            "Committee Name",
            value="Audit Committee",
            help="Display name, e.g. Audit Committee",
        )

    with col2:
        external_committee_id = st.text_input(
            "External Committee ID",
            value="AC001",
            help="ID used by DDM; for now, just pick something unique",
        )

    if st.button("Create Committee"):
        payload = {
            "company_external_id": company_external_id or None,
            "committee_name": committee_name,
            "external_committee_id": external_committee_id or None,
        }
        try:
            resp = requests.post(f"{BACKEND_URL}/committees/", json=payload)
            if resp.status_code == 200:
                st.success(f"Committee created: {resp.json()}")
            else:
                st.error(f"Error {resp.status_code}: {resp.text}")
        except Exception as e:
            st.error(f"Request failed: {e}")

    st.markdown("---")
    st.subheader("Existing Committees")

    try:
        resp = requests.get(f"{BACKEND_URL}/committees/")
        if resp.status_code == 200:
            committees = resp.json()
            if not committees:
                st.info("No committees yet. Create one above.")
            else:
                st.table(committees)
        else:
            st.error(f"Error fetching committees: {resp.status_code} {resp.text}")
    except Exception as e:
        st.error(f"Failed to fetch committees: {e}")

    st.markdown("---")
    st.subheader("Create Meeting")

    committee_options = []
    try:
        resp = requests.get(f"{BACKEND_URL}/committees/")
        if resp.status_code == 200:
            committee_options = resp.json()
    except Exception:
        committee_options = []

    if committee_options:
        committee_labels = [
            f"{c['external_committee_id'] or 'N/A'} – {c['name']}" for c in committee_options
        ]
        idx = st.selectbox(
            "Select Committee",
            range(len(committee_options)),
            format_func=lambda i: committee_labels[i],
        )
        selected_committee = committee_options[idx]
        selected_committee_external_id = selected_committee["external_committee_id"]
    else:
        selected_committee = None
        selected_committee_external_id = None
        st.warning("No committees available for meetings. Create a committee first.")

    col3, col4 = st.columns(2)
    with col3:
        meeting_name = st.text_input(
            "Meeting Name",
            value="Audit Committee Meeting 2025-11-03",
        )
        meeting_date = st.date_input("Meeting Date", value=date.today())
    with col4:
        external_meeting_id = st.text_input(
            "External Meeting ID",
            value="M2025-11-03",
            help="ID used by DDM; for now just choose something unique",
        )

    if st.button("Create Meeting", disabled=selected_committee_external_id is None):
        payload = {
            "external_committee_id": selected_committee_external_id,
            "meeting_name": meeting_name,
            "meeting_date": meeting_date.isoformat() if meeting_date else None,
            "external_meeting_id": external_meeting_id or None,
        }
        try:
            resp = requests.post(f"{BACKEND_URL}/meetings/", json=payload)
            if resp.status_code == 200:
                st.success(f"Meeting created: {resp.json()}")
            else:
                st.error(f"Error {resp.status_code}: {resp.text}")
        except Exception as e:
            st.error(f"Request failed: {e}")

    st.markdown("---")
    st.subheader("Meetings for a Committee")

    if committee_options:
        idx2 = st.selectbox(
            "Choose Committee to list its meetings",
            range(len(committee_options)),
            format_func=lambda i: committee_labels[i],
            key="list_meetings_committee",
        )
        list_committee = committee_options[idx2]
        list_external_committee_id = list_committee["external_committee_id"]

        try:
            resp = requests.get(
                f"{BACKEND_URL}/meetings/",
                params={"committee_external_id": list_external_committee_id},
            )
            if resp.status_code == 200:
                meetings = resp.json()
                if not meetings:
                    st.info("No meetings for this committee yet.")
                else:
                    st.table(meetings)
            else:
                st.error(f"Error fetching meetings: {resp.status_code} {resp.text}")
        except Exception as e:
            st.error(f"Failed to fetch meetings: {e}")
    else:
        st.info("Create a committee above to see meetings.")



with tab_upload:
    st.subheader("Upload Agenda / Minutes / etc.")

    committees = []
    try:
        resp = requests.get(f"{BACKEND_URL}/committees/")
        if resp.status_code == 200:
            committees = resp.json()
    except Exception:
        committees = []

    if not committees:
        st.warning("No committees found. Create a committee in the previous tab first.")
    else:
        committee_labels = [
            f"{c['external_committee_id'] or 'N/A'} – {c['name']}" for c in committees
        ]
        c_idx = st.selectbox(
            "Committee",
            range(len(committees)),
            format_func=lambda i: committee_labels[i],
            key="upload_committee",
        )
        selected_committee = committees[c_idx]
        committee_external_id = selected_committee["external_committee_id"]

        meetings = []
        try:
            resp = requests.get(
                f"{BACKEND_URL}/meetings/",
                params={"committee_external_id": committee_external_id},
            )
            if resp.status_code == 200:
                meetings = resp.json()
        except Exception:
            meetings = []

        meeting_options = ["(None – committee-level doc)"]
        meeting_map: dict[str, str | None] = {"(None – committee-level doc)": None}
        for m in meetings:
            label = f"{m['external_meeting_id'] or 'N/A'} – {m['name']}"
            meeting_options.append(label)
            meeting_map[label] = m["external_meeting_id"]

        meeting_label = st.selectbox("Meeting", meeting_options)
        selected_meeting_external_id = meeting_map[meeting_label]

        doc_type = st.selectbox(
            "Document Type",
            ["Agenda", "DraftMinutes", "FinalMinutes", "CircularResolution", "Extra1", "Extra2"],
        )

        uploaded_file = st.file_uploader(
            "Choose file",
            type=["pdf", "docx", "doc", "txt", "csv", "xlsx", "xls", "pptx", "ppt"],
        )

        external_document_id = st.text_input(
            "External Document ID (optional)",
            help="ID from DDM if available",
        )

        if st.button("Upload Document", type="primary"):
            if not uploaded_file:
                st.error("Please select a file to upload.")
            else:
                files = {
                    "file": (uploaded_file.name, uploaded_file, uploaded_file.type),
                }
                data = {
                    "external_committee_id": committee_external_id,
                    "doc_type": doc_type,
                    "external_meeting_id": selected_meeting_external_id or "",
                    "external_document_id": external_document_id or "",
                }

                try:
                    resp = requests.post(
                        f"{BACKEND_URL}/documents/upload",
                        data=data,
                        files=files,
                    )
                    if resp.status_code == 200:
                        st.success("Document uploaded and registered.")
                        st.json(resp.json())
                    else:
                        st.error(f"Error {resp.status_code}: {resp.text}")
                except Exception as e:
                    st.error(f"Upload failed: {e}")
